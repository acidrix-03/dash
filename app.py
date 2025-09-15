from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify, g
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
import os
import pandas as pd
from io import BytesIO
from flask_babel import Babel
import openpyxl
import shutil
import time
import glob
from datetime import datetime
import pytz

# Set Manila timezone
LOCAL_TIMEZONE = pytz.timezone("Asia/Manila")

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your_secret_key')

# Babel setup for multi-language support
babel = Babel(app)

# UPLOAD FOLDER for file uploads
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Allowed file extensions
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx'}

# Database setup
def init_db():
    try:
        with sqlite3.connect('users.db') as conn:
            conn.execute('''CREATE TABLE IF NOT EXISTS users
                            (id INTEGER PRIMARY KEY AUTOINCREMENT,
                             name TEXT NOT NULL,
                             username TEXT NOT NULL UNIQUE,
                             password TEXT NOT NULL,
                             role TEXT NOT NULL)''')
            admin_password = generate_password_hash('12345', method='pbkdf2:sha256')
            conn.execute('''INSERT OR IGNORE INTO users (name, username, password, role)
                            VALUES ('Admin', 'Admin', ?, 'admin')''', (admin_password,))
            conn.commit()
        print("Database initialized successfully")
    except sqlite3.DatabaseError as db_err:
        print(f"Database2 error: {db_err}")
    except Exception as e:
        print(f"General error: {e}")

    try:
        with sqlite3.connect('documents.db') as conn:
            # Create CTO Application table
            conn.execute('''CREATE TABLE IF NOT EXISTS cto_application (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                position TEXT NOT NULL,
                days INTEGER NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                recommending_approval TEXT DEFAULT NULL, 
                approval_status TEXT DEFAULT 'Pending',
                date_approved TEXT DEFAULT NULL   
            )''')

            # Create a table for recommended applications
            conn.execute('''CREATE TABLE IF NOT EXISTS recommended_applications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                app_id INTEGER NOT NULL,
                app_type TEXT NOT NULL,
                name TEXT NOT NULL,
                position TEXT NOT NULL,
                days INTEGER,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                destination TEXT,
                purpose TEXT,
                leave_type TEXT,
                date_recommended TEXT NOT NULL
            )''')

            # Create a table for approved applications
            conn.execute('''CREATE TABLE IF NOT EXISTS approved_applications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                app_type TEXT NOT NULL,
                name TEXT NOT NULL,
                position TEXT NOT NULL,
                days INTEGER NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                destination TEXT,
                purpose TEXT,
                date_recommended TEXT NOT NULL
            )''')

            # Create Leave Application table
            conn.execute('''CREATE TABLE IF NOT EXISTS leave_application (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                position TEXT NOT NULL,
                days INTEGER NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                recommending_approval TEXT DEFAULT NULL, 
                approval_status TEXT DEFAULT 'Pending',
                date_approved TEXT DEFAULT NULL
            )''')

            # Create Travel Authority table
            conn.execute('''CREATE TABLE IF NOT EXISTS travel_authority (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                position TEXT NOT NULL,
                purpose TEXT NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                destination TEXT NOT NULL,
                recommending_approval TEXT DEFAULT NULL, 
                approval_status TEXT DEFAULT 'Pending',
                date_approved TEXT DEFAULT NULL 
            )''')
        print("Database initialized successfully")
    except sqlite3.DatabaseError as db_err:
        print(f"Database3 error: {db_err}")
    except Exception as e:
        print(f"General error: {e}")

def init_db():
    with sqlite3.connect('chat.db') as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS messages (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            message TEXT NOT NULL
                        )''')
        conn.commit()
import sqlite3


def add_rejection_comment_column():
    with sqlite3.connect('documents.db') as conn:
        # Adding rejection_comment column to cto_application
        conn.execute('ALTER TABLE cto_application ADD COLUMN rejection_comment TEXT;')
        conn.execute('ALTER TABLE leave_application ADD COLUMN rejection_comment TEXT;')
        conn.execute('ALTER TABLE travel_authority ADD COLUMN rejection_comment TEXT;')
        conn.commit()
        print("Columns added successfully.")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Routes
@app.route('/add_column')
def add_column():
    try:
        with sqlite3.connect('documents.db') as conn:
            conn.execute('ALTER TABLE recommended_applications ADD COLUMN recommending_approval TEXT')
        return "Column 'recommending_approval' added successfully!"
    except sqlite3.OperationalError as e:
        return f"Error: {e}"

def get_db_connection():
    conn = sqlite3.connect('users.db')
    conn.row_factory = sqlite3.Row
    return conn

@app.route('/add_recommending_approval_column')
def add_recommending_approval_column():
    try:
        with sqlite3.connect('documents.db') as conn:
            conn.execute('ALTER TABLE recommended_applications ADD COLUMN recommending_approval TEXT')
        return "Column 'recommending_approval' added successfully!"
    except sqlite3.OperationalError as e:
        return f"Error: {e}"

@app.route('/search_user')
def search_user():
    query = request.args.get('query')
    if query:
        with sqlite3.connect('users.db') as conn:
            user = conn.execute("SELECT * FROM users WHERE name = ? OR username = ?", (query, query)).fetchone()
            return jsonify({"exists": bool(user)})
    return jsonify({"exists": False})

@app.route('/search_users', methods=['GET'])
def search_users():
    partial_username = request.args.get('q', '')  # Get the query parameter 'q' for the username
    if partial_username:
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        
        # Query to find usernames that start with the user's input (use `LIKE` for partial matches)
        cursor.execute("SELECT username FROM users WHERE username LIKE ?", (partial_username + '%',))
        matching_usernames = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        return jsonify(matching_usernames)
    return jsonify([])  # Return an empty list if no matches are found

# Function to fetch office suggestions from the database
def fetch_offices(query):
    connection = sqlite3.connect('users.db')
    cursor = connection.cursor()
    
    # Fetch offices matching the query
    cursor.execute("SELECT DISTINCT office FROM users WHERE office LIKE ?", ('%' + query + '%',))
    offices = [row[0] for row in cursor.fetchall()]
    
    connection.close()
    return offices

@app.route('/search_offices')
def search_offices():
    q = request.args.get('q', '')
    offices = fetch_offices(q)
    return jsonify(offices)

@app.route('/')
def index():
    return render_template('login.html')

# Track logged-in users
logged_in_users = set()

@app.route('/login', methods=['POST'])
def login():
    name = request.form['username']
    password = request.form['password']
    with sqlite3.connect('users.db') as conn:
        user = conn.execute('SELECT * FROM users WHERE name = ?', (name,)).fetchone()
        if user and check_password_hash(user[3], password):
            session['user_id'] = user[0]
            session['username'] = user[2]
            session['position'] = user[5]
            session['role'] = user[4]
            session['division'] = user[8]
            session['office'] = user[6]
            session['salary'] = user[7]
            logged_in_users.add(user[0])  # Add user to logged-in users set
            return redirect(url_for('user_dashboard'))
        else:
            flash('Invalid credentials')
            return redirect(url_for('index'))

@app.route('/logout')
def logout():
    user_id = session.get('user_id')
    if user_id in logged_in_users:
        logged_in_users.remove(user_id)  # Remove user from logged-in users set
    session.clear()
    return redirect(url_for('index'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name']
        username = request.form['username']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        
        if password != confirm_password:
            flash('Passwords do not match')
            return redirect(url_for('register'))

        # Check if the name already exists in the database
        with sqlite3.connect('users.db') as conn:
            user = conn.execute('SELECT * FROM users WHERE name = ?', (name,)).fetchone()
            if user:
                flash('Name already exists')
                return redirect(url_for('register'))
            
            # Proceed with registration if name is unique
            try:
                conn.execute('INSERT INTO users (name, username, password, role) VALUES (?, ?, ?, ?)', 
                             (name, username, generate_password_hash(password, method='pbkdf2:sha256'), 'user'))
                conn.commit()
                flash('Registration successful')
                return redirect(url_for('index'))
            except sqlite3.IntegrityError:
                flash('Username already exists')
                return redirect(url_for('register'))
    return render_template('register.html')

@app.route('/submit_document', methods=['GET', 'POST'])
def submit_document():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))
    if request.method == 'POST':
        name = request.form['name']
        division = request.form['division']
        document = request.files['document']
        submission_date = datetime.now() 
        if document and allowed_file(document.filename):
            filename = secure_filename(document.filename)
            document.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            flash('Document uploaded successfully')
        else:
            flash('Invalid file format')
        return redirect(url_for('submit_document'))
    return render_template('submit_document.html')


@app.route('/admin_dashboard')
def admin_dashboard():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied')
        return redirect(url_for('index'))

    with sqlite3.connect('documents.db') as conn:
        # Count all applications
        total_cto_applications = conn.execute('SELECT COUNT(*) FROM cto_application').fetchone()[0]
        total_leave_applications = conn.execute('SELECT COUNT(*) FROM leave_application').fetchone()[0]
        total_travel_authorities = conn.execute('SELECT COUNT(*) FROM travel_authority').fetchone()[0]
        
        # Count pending applications
        pending_cto_applications = conn.execute('SELECT COUNT(*) FROM cto_application WHERE approval_status = "Pending"').fetchone()[0]
        pending_leave_applications = conn.execute('SELECT COUNT(*) FROM leave_application WHERE approval_status = "Pending"').fetchone()[0]
        pending_travel_authorities = conn.execute('SELECT COUNT(*) FROM travel_authority WHERE approval_status = "Pending"').fetchone()[0]

        # Count applications for recommending
        recommending_cto_applications = conn.execute('SELECT COUNT(*) FROM cto_application WHERE recommending_approval IS NOT NULL AND approval_status = "Pending"').fetchone()[0]
        recommending_leave_applications = conn.execute('SELECT COUNT(*) FROM leave_application WHERE recommending_approval IS NOT NULL AND approval_status = "Pending"').fetchone()[0]
        recommending_travel_authorities = conn.execute('SELECT COUNT(*) FROM travel_authority WHERE recommending_approval IS NOT NULL AND approval_status = "Pending"').fetchone()[0]

        # Count approved applications
        approved_cto_applications = conn.execute('SELECT COUNT(*) FROM cto_application WHERE approval_status = "Approved"').fetchone()[0]
        approved_leave_applications = conn.execute('SELECT COUNT(*) FROM leave_application WHERE approval_status = "Approved"').fetchone()[0]
        approved_travel_authorities = conn.execute('SELECT COUNT(*) FROM travel_authority WHERE approval_status = "Approved"').fetchone()[0]

    with sqlite3.connect('document_tracker.db') as conn:
        # Count documents
        total_submitted = conn.execute('SELECT COUNT(*) FROM documents').fetchone()[0]
        total_forwarded = conn.execute('SELECT COUNT(*) FROM forwarding_history').fetchone()[0]
        total_received = conn.execute('SELECT COUNT(*) FROM receiving_history').fetchone()[0]

    # Prepare the list of currently logged-in usernames
    global logged_in_users, logged_in_usernames
    if 'logged_in_users' not in globals():
        logged_in_users = set()
    if 'logged_in_usernames' not in globals():
        logged_in_usernames = set()
    usernames_list = list(logged_in_usernames) if 'logged_in_usernames' in globals() else []

    # Pass all stats as a list of dicts for a single-row table
    stats_row = [
        {
            "label": "CTO Applications",
            "total": total_cto_applications,
            "pending": pending_cto_applications,
            "recommending": recommending_cto_applications,
            "approved": approved_cto_applications
        },
        {
            "label": "Leave Applications",
            "total": total_leave_applications,
            "pending": pending_leave_applications,
            "recommending": recommending_leave_applications,
            "approved": approved_leave_applications
        },
        {
            "label": "Travel Authorities",
            "total": total_travel_authorities,
            "pending": pending_travel_authorities,
            "recommending": recommending_travel_authorities,
            "approved": approved_travel_authorities
        },
        {
            "label": "Documents",
            "total": total_submitted,
            "forwarded": total_forwarded,
            "received": total_received
        }
    ]

    return render_template(
        'admin_dashboard.html',
        stats_row=stats_row,
        logged_in_users=len(logged_in_users),
        logged_in_usernames=usernames_list
    )

@app.route('/view_users')
def view_users():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied')
        return redirect(url_for('index'))

    search = request.args.get('search', '')
    letter = request.args.get('letter', '')
    page = request.args.get('page', 1, type=int)  # Current page number (default is 1)
    per_page = 50

    query = 'SELECT id, name, username, position, role, office, division FROM users WHERE 1=1'
    params = []

    if search:
        query += ' AND (name LIKE ? OR username LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])

    if letter:
        query += ' AND (name LIKE ? OR username LIKE ?)'
        params.extend([f'{letter}%', f'{letter}%'])

    query += ' LIMIT ? OFFSET ?'
    params.extend([per_page, (page - 1) * per_page])

    with sqlite3.connect('users.db') as conn:
        total_users = conn.execute('SELECT COUNT(*) FROM users WHERE 1=1').fetchone()[0]
        users = conn.execute(query, params).fetchall()

    # Calculate total pages
    total_pages = (total_users + per_page - 1) // per_page

    return render_template('view_users.html', users=users, page=page, total_pages=total_pages, search=search, letter=letter)

@app.route('/clear_data', methods=['POST'])
def clear_data():
    admin_password = request.form['admin_password']
    with sqlite3.connect('users.db') as conn:
        admin_user = conn.execute('SELECT * FROM users WHERE username = ?', ('Admin',)).fetchone()
        if admin_user and check_password_hash(admin_user[3], admin_password):
            # Clear data from documents.db
            with sqlite3.connect('documents.db') as doc_conn:
                # Clear all application data
                doc_conn.execute('DELETE FROM cto_application')
                doc_conn.execute('DELETE FROM leave_application')
                doc_conn.execute('DELETE FROM travel_authority')
                doc_conn.execute('DELETE FROM recommended_applications')
                doc_conn.execute('DELETE FROM approved_applications')

            # Clear data from document_tracker.db
            with sqlite3.connect('document_tracker.db') as tracker_conn:
                tracker_conn.execute('DELETE FROM documents')
                tracker_conn.execute('DELETE FROM forwarding_history')
                tracker_conn.commit()
            
            # Clear all generated Excel files in the static/generated_files folder
            output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'generated_files')
            if os.path.exists(output_dir):
                # Delete all files in the output directory
                for filename in os.listdir(output_dir):
                    file_path = os.path.join(output_dir, filename)
                    try:
                        if os.path.isfile(file_path):
                            os.remove(file_path)  # Delete the file
                    except Exception as e:
                        print(f"Error deleting file {file_path}: {e}")
            
            flash('All data and generated files have been cleared successfully', 'success')
        else:
            flash('Invalid admin password', 'danger')
    
    return redirect(url_for('admin_dashboard'))


@app.route('/change_position', methods=['GET', 'POST'])
def change_position():
    if 'user_id' not in session:
        flash('Please log in first.')
        return redirect(url_for('index'))

    if request.method == 'POST':
        new_position = request.form['position']
        
        with sqlite3.connect('users.db') as conn:
            conn.execute('UPDATE users SET position = ? WHERE id = ?', (new_position, session['user_id']))
            conn.commit()
        
        session['position'] = new_position  # Update session with new position
        flash('Position updated successfully.')
        return redirect(url_for('user_dashboard'))

    # Check if 'position' exists in session
    current_position = session.get('position', 'Not set')  # Use default value if not set
    return render_template('change_position.html', current_position=current_position)

@app.route('/export_excel')
def export_excel():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied')
        return redirect(url_for('index'))

    with sqlite3.connect('documents.db') as conn:
        travel_authorities = pd.read_sql_query('SELECT * FROM travel_authority', conn)
        cto_applications = pd.read_sql_query('SELECT * FROM cto_application', conn)
        leave_applications = pd.read_sql_query('SELECT * FROM leave_application', conn)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        travel_authorities.to_excel(writer, sheet_name='Travel Authorities', index=False)
        cto_applications.to_excel(writer, sheet_name='CTO Applications', index=False)
        leave_applications.to_excel(writer, sheet_name='Leave Applications', index=False)

    output.seek(0)
    return send_file(output, download_name='admin_data.xlsx', as_attachment=True)


@app.route('/export_users_excel')
def export_users_excel():
    with sqlite3.connect('users.db') as conn:
        users_df = pd.read_sql_query("SELECT * FROM users", conn)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        users_df.to_excel(writer, sheet_name='Users', index=False)
    
    output.seek(0)
    return send_file(output, download_name='users_data.xlsx', as_attachment=True)

@app.route('/import_users_excel', methods=['POST'])
def import_users_excel():
    file = request.files['file']  # Handle file upload from the form
    if not file:
        flash('No file selected', 'error')
        return redirect(url_for('admin_dashboard'))

    try:
        users_df = pd.read_excel(file)
        
        # Convert all passwords to strings before hashing
        users_df['password'] = users_df['password'].astype(str)
        users_df['password'] = users_df['password'].apply(lambda x: generate_password_hash(x, method='pbkdf2:sha256'))

        with sqlite3.connect('users.db') as conn:
            for _, row in users_df.iterrows():
                try:
                    conn.execute('INSERT INTO users (name, username, password, role, position, salary) VALUES (?, ?, ?, ?, ?, ?)', 
                                 (row['name'], row['username'], row['password'], row['role'], row['position'], row['salary']))
                except sqlite3.IntegrityError:
                    flash(f"User {row['username']} already exists", 'warning')
        
        flash('Users imported successfully', 'success')
        return redirect(url_for('admin_dashboard'))
    except Exception as e:
        flash(f"Error importing users: {e}", 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/cancel_application/<int:app_id>/<string:app_type>', methods=['POST'])
def cancel_application(app_id, app_type):
    if 'user_id' not in session:
        return jsonify({'error': 'Access denied'}), 403

    try:
        with sqlite3.connect('documents.db') as conn:
            if app_type == 'cto':
                conn.execute('DELETE FROM cto_application WHERE id = ?', (app_id,))
            elif app_type == 'leave':
                conn.execute('DELETE FROM leave_application WHERE id = ?', (app_id,))
            elif app_type == 'travel':
                conn.execute('DELETE FROM travel_authority WHERE id = ?', (app_id,))
            conn.commit()
        return jsonify({'success': 'Application cancelled successfully'})
    except Exception as e:
        print(f"Error occurred: {e}")
        return jsonify({'error': 'Failed to cancel application'}), 500



@app.route('/reject_application/<int:app_id>', methods=['POST'])
def reject_application(app_id):
    if 'user_id' not in session or session.get('role') != 'approver':
        return jsonify({'error': 'Access denied'}), 403

    application_type = request.form['application_type']
    rejection_comment = request.form['rejection_comment']  # Capture the rejection comment

    with sqlite3.connect('documents.db') as conn:
        # Fetch the application data and update the status based on the type
        if application_type == 'cto':
            application = conn.execute('SELECT name, position, days, start_date, end_date FROM cto_application WHERE id = ?', (app_id,)).fetchone()

            name, position, days, start_date, end_date = application
            conn.execute('UPDATE cto_application SET approval_status = "Rejected", rejection_comment = ? WHERE id = ?', (rejection_comment, app_id))
            conn.execute('''INSERT INTO rejected_applications 
                            (app_type, name, position, days, start_date, end_date, rejection_comment, date_rejected)
                            VALUES (?, ?, ?, ?, ?, ?, ?, date("now"))''', 
                            (application_type, name, position, days, start_date, end_date, rejection_comment))

        elif application_type == 'leave':
            application = conn.execute('SELECT name, position, days, start_date, end_date FROM leave_application WHERE id = ?', (app_id,)).fetchone()
            name, position, days, start_date, end_date = application
            conn.execute('UPDATE leave_application SET approval_status = "Rejected", rejection_comment = ? WHERE id = ?', (rejection_comment, app_id))
            conn.execute('''INSERT INTO rejected_applications 
                            (app_type, name, position, days, start_date, end_date, rejection_comment, date_rejected)
                            VALUES (?, ?, ?, ?, ?, ?, ?, date("now"))''', 
                            (application_type, name, position, days, start_date, end_date, rejection_comment))

        elif application_type == 'travel_authority':
            application = conn.execute('SELECT name, position, purpose, start_date, end_date, destination FROM travel_authority WHERE id = ?', (app_id,)).fetchone()
            name, position, purpose, start_date, end_date, destination = application
            conn.execute('UPDATE travel_authority SET approval_status = "Rejected", rejection_comment = ? WHERE id = ?', (rejection_comment, app_id))
            conn.execute('''INSERT INTO rejected_applications 
                            (app_type, name, position, purpose, start_date, end_date, destination, rejection_comment, date_rejected)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, date("now"))''', 
                            (application_type, name, position, purpose, start_date, end_date, destination, rejection_comment))

        conn.commit()

    return jsonify({'success': True})


@app.route('/cto_application', methods=['GET', 'POST'])
def cto_application():
    user_id = session['user_id']  # Get the logged-in user's ID

    if request.method == 'POST':
        # Handle form submission
        name = request.form['name']
        position = request.form['position']
        days = request.form['days']
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        recommending_approval = request.form['recommending_approval']

        # Insert the new CTO application into the database
        with sqlite3.connect('documents.db') as conn:
            conn.execute('''INSERT INTO cto_application (name, position, days, start_date, end_date, user_id, recommending_approval)
                            VALUES (?, ?, ?, ?, ?, ?, ?)''', (name, position, days, start_date, end_date, user_id, recommending_approval))
            conn.commit()

        flash('CTO Application submitted successfully!')
        return redirect(url_for('user_dashboard'))

    else:
        # Fetch existing CTO application for this user
        with sqlite3.connect('documents.db') as conn:
            conn.row_factory = sqlite3.Row  # Allows access to rows as dictionaries
            cto_application = conn.execute('SELECT * FROM cto_application WHERE user_id = ?', (user_id,)).fetchone()

        # Fetch Unit Heads and Recommenders for the dropdown
        with sqlite3.connect('users.db') as conn:
            approving_users = [row[0] for row in conn.execute('SELECT username FROM users WHERE role IN ("unit_head", "recommender")').fetchall()]

        return render_template('cto_application.html', approving_users=approving_users, cto_application=cto_application)


@app.route('/submit_and_print_cto_application_excel', methods=['POST'])
def submit_and_print_cto_application_excel():
    # Get form data
    name = request.form['name']
    position = request.form['position']
    days = request.form['days']
    start_date = request.form['start_date']
    end_date = request.form['end_date']
    recommender = request.form['recommending_approval']

    user_id = session['user_id']  # Get the logged-in user's ID

    # Insert the new CTO application into the database
    with sqlite3.connect('documents.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO cto_application (name, position, days, start_date, end_date, user_id, recommending_approval)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (name, position, days, start_date, end_date, user_id, recommender))
        app_id = cursor.lastrowid  # Get the last inserted ID
        conn.commit()

    # Fetch recommender's position from the users.db
    with sqlite3.connect('users.db') as user_conn:
        recommender_position = user_conn.execute('SELECT position FROM users WHERE username = ?', (recommender,)).fetchone()
        recommender_position = recommender_position[0] if recommender_position else "Unknown Position"

    # # Path to the template file in static folder
    # template_path = os.path.join('static', 'cto_application_template.xlsx')

    # # Load the workbook and fill in the data
    # wb = openpyxl.load_workbook(template_path)
    # sheet = wb.active
    # sheet['I19'] = name  
    # sheet['I20'] = position  
    # sheet['C11'] = start_date  
    # sheet['E11'] = end_date  
    # sheet['E19'] = recommender  
    # sheet['E20'] = recommender_position  
    # submission_date = datetime.now().strftime('%m-%d-%Y')
    # sheet['J22'] = submission_date

    # # Ensure the 'generated_files' directory exists
    # output_directory = os.path.join('static', 'generated_files')
    # if not os.path.exists(output_directory):
    #     os.makedirs(output_directory)

    # # Save the file with a unique name (to avoid overwriting) cto_save
    # output_filename = f'cto_application_{app_id}_{int(time.time())}.xlsx'
    # output_path = os.path.join(output_directory, output_filename)
    # wb.save(output_path)
    # wb.close()

    flash('CTO Application Submitted Successfully!', 'success')
    return redirect(url_for('user_dashboard'))
    # # Redirect to the download route with the generated file's name
    # return redirect(url_for('download_cto_application', filename=output_filename))



@app.route('/download_cto_application/<filename>', methods=['GET'])
def download_cto_application(filename):
    file_path = os.path.join('static', 'generated_files', filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        flash('File not found.', 'danger')
        return redirect(url_for('user_dashboard'))


@app.route('/leave_application', methods=['GET', 'POST'])
def leave_application():
    user_id = session['user_id']  # Get the logged-in user's ID

    if request.method == 'POST':
        # Handle form submission
        name = request.form['name']
        position = request.form['position']
        days = request.form['days']
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        recommending_approval = request.form['recommending_approval']

        # Insert the new leave application into the database
        with sqlite3.connect('documents.db') as conn:
            conn.execute('''INSERT INTO leave_application (name, position, days, start_date, end_date, user_id, recommending_approval)
                            VALUES (?, ?, ?, ?, ?, ?, ?)''', (name, position, days, start_date, end_date, user_id, recommending_approval))
            conn.commit()

        flash('Leave Application submitted successfully!')
        return redirect(url_for('user_dashboard'))

    else:
        # Fetch user details (office and salary) for this user from users.db
        with sqlite3.connect('users.db') as conn:
            conn.row_factory = sqlite3.Row  # Allows access to rows as dictionaries
            user_details = conn.execute('SELECT office, salary FROM users WHERE id = ?', (user_id,)).fetchone()

        # Check if user_details is None (user not found)
        if user_details is None:
            flash('User details not found.', 'danger')
            return redirect(url_for('user_dashboard'))

        # Fetch Recommenders for the dropdown
        with sqlite3.connect('users.db') as conn:
            approving_users = [row[0] for row in conn.execute('SELECT username FROM users WHERE role IN ("unit_head", "recommender")').fetchall()]

        return render_template('leave_application.html', approving_users=approving_users, user_details=user_details)

@app.route('/submit_and_print_leave_application_excel', methods=['POST'])
def submit_and_print_leave_application_excel():
    # Get form data
    name = request.form['name']
    position = request.form['position']
    office = request.form['office']  
    salary = request.form['salary']  
    days = request.form['days']
    start_date = request.form['start_date']
    end_date = request.form['end_date']
    leave_type = request.form['leave_type']
    recommender = request.form['recommending_approval']

    user_id = session['user_id']  # Get the logged-in user's ID

    # Insert the new Leave application into the database
    with sqlite3.connect('documents.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''INSERT INTO leave_application (name, position, days, start_date, end_date, leave_type, user_id, recommending_approval)
                          VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', 
                       (name, position, days, start_date, end_date, leave_type, user_id, recommender))
        app_id = cursor.lastrowid  # Get the last inserted ID
        conn.commit()

    # Fetch recommender's position from the users.db
    with sqlite3.connect('users.db') as user_conn:
        recommender_position = user_conn.execute('SELECT position FROM users WHERE username = ?', (recommender,)).fetchone()
        recommender_position = recommender_position[0] if recommender_position else "Unknown Position"

    # Path to the template file in static folder
    template_path = os.path.join('static', 'leave_application_template.xlsx')

    # Load the workbook and fill in the data
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    sheet['E5'] = name  
    sheet['E7'] = position
    sheet['C5'] = office
    sheet['I7'] = salary
    sheet['C45'] = days
    sheet['C48'] = start_date
    sheet['D48'] = end_date
    sheet['I59'] = recommender
    submission_date = datetime.now().strftime('%m-%d-%Y')
    sheet['C7'] = submission_date

    # # Ensure the 'generated_files' directory exists
    # output_directory = os.path.join('static', 'generated_files')
    # if not os.path.exists(output_directory):
    #     os.makedirs(output_directory)

    # # Save the file with a unique name (to avoid overwriting) leave_save
    # output_filename = f'leave_application_{app_id}_{int(time.time())}.xlsx'
    # output_path = os.path.join(output_directory, output_filename)
    # wb.save(output_path)
    # wb.close()

    flash('Leave Application Submitted Successfully!', 'success')
    return redirect(url_for('user_dashboard'))
    # Redirect to the download route with the generated file's name
    # return redirect(url_for('download_leave_application', filename=output_filename))


@app.route('/download_leave_application/<filename>', methods=['GET'])
def download_leave_application(filename):
    file_path = os.path.join('static', 'generated_files', filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        flash('File not found.', 'danger')
        return redirect(url_for('user_dashboard'))


@app.route('/travel_authority', methods=['GET', 'POST'])
def travel_authority():
    user_id = session['user_id']  # Get the logged-in user's ID
    
    if request.method == 'POST':
        # Handle form submission (not required here as it redirects to another route)
        return redirect(url_for('submit_and_print_travel_authority_excel'))
    
    else:
        # Fetch recommenders for dropdown
        with sqlite3.connect('users.db') as conn:
            approving_users = [row[0] for row in conn.execute('SELECT username FROM users WHERE role IN ("unit_head", "recommender")').fetchall()]
        return render_template('travel_authority.html', approving_users=approving_users)

@app.route('/submit_and_print_travel_authority_excel', methods=['POST'])
def submit_and_print_travel_authority_excel():
    # Get form data
    name = request.form['name']
    position = request.form['position']
    purpose = request.form['purpose']
    host = request.form['host']
    destination = request.form['destination']
    start_date = request.form['start_date']
    end_date = request.form['end_date']
    recommending_approval = request.form['recommending_approval']

    user_id = session['user_id']  # Get the logged-in user's ID

    # Insert the new Travel Authority into the database
    with sqlite3.connect('documents.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''INSERT INTO travel_authority 
                          (name, position, purpose, host, destination, start_date, end_date, user_id, recommending_approval)
                          VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                       (name, position, purpose, host, destination, start_date, end_date, user_id, recommending_approval))
        app_id = cursor.lastrowid  # Get the last inserted ID
        conn.commit()

    # Fetch recommender's position from the users.db
    with sqlite3.connect('users.db') as user_conn:
        recommender_position = user_conn.execute('SELECT position FROM users WHERE username = ?', (recommender,)).fetchone()
        recommender_position = recommender_position[0] if recommender_position else "Unknown Position"

    # Path to the template file in static folder
    template_path = os.path.join('static', 'travel_authority_template.xlsx')

    # Load the workbook and fill in the data
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    sheet['B4'] = name
    sheet['A13'] = name
    sheet['B5'] = position
    sheet['B7'] = purpose
    sheet['B8'] = host
    sheet['B9'] = start_date
    sheet['D9'] = end_date
    sheet['B10'] = destination

    submission_date = datetime.now().strftime('%m-%d-%Y')
    sheet['D14'] = submission_date

    # # Ensure the 'generated_files' directory exists
    # output_directory = os.path.join('static', 'generated_files')
    # if not os.path.exists(output_directory):
    #     os.makedirs(output_directory)

    # # Save the file with a unique name (to avoid overwriting) travel_save
    # output_filename = f'travel_authority_{app_id}_{int(time.time())}.xlsx'
    # output_path = os.path.join(output_directory, output_filename)
    # wb.save(output_path)
    # wb.close()

    flash('Travel Authority Submitted and Excel Generated Successfully!', 'success')
    return redirect(url_for('user_dashboard'))
    # Redirect to the download route with the generated file's name
    # return redirect(url_for('download_travel_application', filename=output_filename))

@app.route('/download_travel_application/<filename>', methods=['GET'])
def download_travel_application(filename):
    file_path = os.path.join('static', 'generated_files', filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        flash('File not found.', 'danger')
        return redirect(url_for('user_dashboard'))

@app.route('/change_role/<int:user_id>', methods=['POST'])
def change_role(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied')
        return redirect(url_for('index'))

    new_role = request.form['role']
    with sqlite3.connect('users.db') as conn:
        conn.execute('UPDATE users SET role = ? WHERE id = ?', (new_role, user_id))
        conn.commit()

    flash('User role updated successfully', 'success')
    return redirect(url_for('view_users'))

@app.route('/change_password/<int:user_id>', methods=['GET', 'POST'])
def change_password(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied')
        return redirect(url_for('index'))

    if request.method == 'POST':
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        if new_password != confirm_password:
            flash('Passwords do not match')
            return redirect(url_for('change_password', user_id=user_id))

        hashed_password = generate_password_hash(new_password, method='pbkdf2:sha256')
        with sqlite3.connect('users.db') as conn:
            conn.execute('UPDATE users SET password = ? WHERE id = ?', (hashed_password, user_id))
            conn.commit()
        flash('Password updated successfully', 'success')
        return redirect(url_for('view_users'))

    return render_template('change_password.html', user_id=user_id)

@app.route('/delete_user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied')
        return redirect(url_for('index'))

    try:
        with sqlite3.connect('users.db') as user_conn:
            user = user_conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
            if user:
                user_conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
                user_conn.commit()

        with sqlite3.connect('documents.db') as doc_conn:
            doc_conn.execute('DELETE FROM travel_authority WHERE user_id = ?', (user_id,))
            doc_conn.execute('DELETE FROM cto_application WHERE user_id = ?', (user_id,))
            doc_conn.execute('DELETE FROM leave_application WHERE user_id = ?', (user_id,))
            doc_conn.commit()

        flash('User and all associated data deleted successfully', 'success')
    except sqlite3.Error as e:
        flash(f"An error occurred: {e}", 'danger')

    return redirect(url_for('view_users'))

@app.route('/change_password_user', methods=['GET', 'POST'])
def change_password_user():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        if new_password != confirm_password:
            flash('New passwords do not match')
            return redirect(url_for('change_password_user'))

        with sqlite3.connect('users.db') as conn:
            user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
            if user and check_password_hash(user[3], current_password):
                hashed_password = generate_password_hash(new_password, method='pbkdf2:sha256')
                conn.execute('UPDATE users SET password = ? WHERE id = ?', (hashed_password, session['user_id']))
                conn.commit()
                flash('Password updated successfully', 'success')
                return redirect(url_for('user_dashboard'))
            else:
                flash('Current password is incorrect')
                return redirect(url_for('change_password_user'))

    return render_template('change_password_user.html')

@app.route('/applications_dashboard', methods=['GET', 'POST'])
def applications_dashboard():
    user_id = session['user_id']  # Get the logged-in user's ID
    
    if request.method == 'POST':
        # Retrieve form data
        application_type = request.form['application_type']  # 'cto', 'leave', or 'travel'
        name = request.form['name']
        position = request.form['position']
        days = request.form.get('days')
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        recommending_approval = request.form['recommending_approval']
        
        # Create a DataFrame based on the form data
        data = {
            'Name': [name],
            'Position': [position],
            'Days Applied': [days] if days else '',
            'Start Date': [start_date],
            'End Date': [end_date],
            'Recommender': [recommending_approval]
        }

        # Additional fields for leave or travel applications
        if application_type == 'leave':
            leave_type = request.form['leave_type']
            data['Leave Type'] = [leave_type]
            flash('Leave Application submitted successfully!')
        elif application_type == 'travel':
            purpose = request.form['purpose']
            destination = request.form['destination']
            data['Purpose'] = [purpose]
            data['Destination'] = [destination]
            flash('Travel Authority submitted successfully!')
        else:
            flash('CTO Application submitted successfully!')

        # Insert application into the database
        with sqlite3.connect('documents.db') as conn:
            if application_type == 'cto':
                conn.execute('''INSERT INTO cto_application (name, position, days, start_date, end_date, user_id, recommending_approval)
                                VALUES (?, ?, ?, ?, ?, ?, ?)''', (name, position, days, start_date, end_date, user_id, recommending_approval))
            elif application_type == 'leave':
                conn.execute('''INSERT INTO leave_application (name, position, days, start_date, end_date, user_id, recommending_approval)
                                VALUES (?, ?, ?, ?, ?, ?, ?)''', (name, position, days, start_date, end_date, user_id, recommending_approval))
            elif application_type == 'travel':
                conn.execute('''INSERT INTO travel_authority (name, position, purpose, start_date, end_date, destination, user_id, recommending_approval)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', (name, position, purpose, start_date, end_date, destination, user_id, recommending_approval))
            conn.commit()

        # Create an Excel file in memory
        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=application_type.capitalize())
        
        output.seek(0)  # Reset the pointer to the start of the stream

        # Send the file as a download
        filename = f"{application_type}_application_{name}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    else:
        # Fetch required data for GET request
        with sqlite3.connect('documents.db') as conn:
            conn.row_factory = sqlite3.Row
            cto_application = conn.execute('SELECT * FROM cto_application WHERE user_id = ?', (user_id,)).fetchone()
            leave_application = conn.execute('SELECT * FROM leave_application WHERE user_id = ?', (user_id,)).fetchone()
            travel_authority = conn.execute('SELECT * FROM travel_authority WHERE user_id = ?', (user_id,)).fetchone()

        with sqlite3.connect('users.db') as conn:
            approving_users = [row[0] for row in conn.execute('SELECT username FROM users WHERE role IN ("unit_head", "recommender")').fetchall()]
            user_details = conn.execute('SELECT office, salary FROM users WHERE id = ?', (user_id,)).fetchone()
            if user_details is None:
                flash('User details not found.', 'danger')
                return redirect(url_for('user_dashboard'))

        return render_template(
            'applications_dashboard.html', 
            approving_users=approving_users, 
            user_details=user_details, 
            cto_application=cto_application,
            leave_application=leave_application,
            travel_authority=travel_authority
        )


@app.route('/approver_dashboard')
def approver_dashboard():
    if 'user_id' not in session or session.get('role') != 'approver':
        flash('Access denied')
        return redirect(url_for('index'))

    approver_username = session['username']  # Correctly fetch the logged-in approver's username

    # Fetch applications assigned for approval
    with sqlite3.connect('documents.db') as conn:
        cto_applications = conn.execute('SELECT *, recommend_name FROM cto_application WHERE recommending_approval = "Recommended" AND approval_status = "Pending"').fetchall()
        leave_applications = conn.execute('SELECT *, recommend_name FROM leave_application WHERE recommending_approval = "Recommended" AND approval_status = "Pending"').fetchall()
        travel_authorities = conn.execute('SELECT *, recommend_name FROM travel_authority WHERE recommending_approval = "Recommended" AND approval_status = "Pending"').fetchall()

    return render_template(
        'approver_dashboard.html',
        cto_applications=cto_applications,
        leave_applications=leave_applications,
        travel_authorities=travel_authorities
    )



@app.route('/approve_application/<int:app_id>', methods=['POST'])
def approve_application(app_id):
    if 'user_id' not in session or session.get('role') != 'approver':
        return jsonify({'error': 'Access denied'}), 403

    application_type = request.form['application_type']
    
    with sqlite3.connect('documents.db') as conn:
        # Fetch the application data based on the type
        if application_type == 'cto':
            application = conn.execute('SELECT name, position, days, start_date, end_date FROM cto_application WHERE id = ?', (app_id,)).fetchone()
            name, position, days, start_date, end_date = application
            conn.execute('UPDATE cto_application SET approval_status = "Approved" WHERE id = ?', (app_id,))
            conn.execute('''INSERT INTO approved_applications 
                            (app_type, name, position, days, start_date, end_date, destination, purpose, date_recommended)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, date("now"))''', 
                            (application_type, name, position, days, start_date, end_date, None, None))
        
        elif application_type == 'leave':
            application = conn.execute('SELECT name, position, days, start_date, end_date FROM leave_application WHERE id = ?', (app_id,)).fetchone()
            name, position, days, start_date, end_date = application
            conn.execute('UPDATE leave_application SET approval_status = "Approved" WHERE id = ?', (app_id,))
            conn.execute('''INSERT INTO approved_applications 
                            (app_type, name, position, days, start_date, end_date, destination, purpose, date_recommended)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, date("now"))''', 
                            (application_type, name, position, days, start_date, end_date, None, None))
        
        elif application_type == 'travel_authority':
            application = conn.execute('SELECT name, position, purpose, start_date, end_date, destination FROM travel_authority WHERE id = ?', (app_id,)).fetchone()
            name, position, purpose, start_date, end_date, destination = application
            conn.execute('UPDATE travel_authority SET approval_status = "Approved" WHERE id = ?', (app_id,))
            conn.execute('''INSERT INTO approved_applications 
                            (app_type, name, position, days, start_date, end_date, destination, purpose, date_recommended)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, date("now"))''', 
                            (application_type, name, position, 0, start_date, end_date, destination, purpose))
        
        conn.commit()
    
    return jsonify({'success': True})

@app.route('/recommend_approval/<int:app_id>', methods=['POST'])
def recommend_approval(app_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Access denied'}), 403

    user_role = session.get('role')
    username = session.get('username')

    # Allow both recommenders and unit heads to recommend approval
    if user_role not in ['recommender', 'unit_head']:
        return jsonify({'error': 'Access denied'}), 403

    application_type = request.form['application_type']

    with sqlite3.connect('documents.db') as conn:
        if application_type == 'cto':
            application = conn.execute('SELECT * FROM cto_application WHERE id = ?', (app_id,)).fetchone()
            conn.execute('''
                UPDATE cto_application
                SET recommending_approval = "Recommended", recommend_name = ?
                WHERE id = ?
            ''', (username, app_id))
        elif application_type == 'leave':
            application = conn.execute('SELECT * FROM leave_application WHERE id = ?', (app_id,)).fetchone()
            conn.execute('''
                UPDATE leave_application
                SET recommending_approval = "Recommended", recommend_name = ?
                WHERE id = ?
            ''', (username, app_id))
        elif application_type == 'travel_authority':
            application = conn.execute('SELECT * FROM travel_authority WHERE id = ?', (app_id,)).fetchone()
            conn.execute('''
                UPDATE travel_authority
                SET recommending_approval = "Recommended", recommend_name = ?
                WHERE id = ?
            ''', (username, app_id))
        else:
            return jsonify({'error': 'Invalid application type'}), 400

        # Insert into recommended_applications table
        conn.execute('''
            INSERT INTO recommended_applications 
            (app_id, app_type, name, position, days, start_date, end_date, destination, purpose, leave_type, date_recommended, recommending_approval)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, date("now"), ?)
        ''', (
            app_id,  # Application ID
            application_type,  # Type of the application (cto/leave/travel_authority)
            application[1],  # Name
            application[2],  # Position
            application[3],  # Days (for leave or CTO)
            application[4],  # Start Date
            application[5],  # End Date
            application[6] if application_type == 'travel_authority' else None,  # Destination for travel authority
            application[7] if application_type == 'travel_authority' else None,  # Purpose for travel authority
            application[6] if application_type == 'leave' else None,  # Leave type for leave applications
            username  # The person recommending approval (e.g., the unit head or recommender)
        ))
        conn.commit()

    return jsonify({'success': True})

def fetch_approver_dashboard_data():
    conn = sqlite3.connect('your_database.db')
    cursor = conn.cursor()

    # Fetch CTO Applications
    cursor.execute('''
        SELECT id, name, position, days, start_date, end_date, recommending_approval
        FROM cto_applications
    ''')
    cto_applications = cursor.fetchall()

    # Fetch Leave Applications
    cursor.execute('''
        SELECT id, name, position, days, start_date, end_date, leave_type, status, recommending_approval
        FROM leave_applications
    ''')
    leave_applications = cursor.fetchall()

    # Fetch Travel Authorities
    cursor.execute('''
        SELECT id, name, position, purpose, destination, start_date, end_date, status, recommending_approval
        FROM travel_authorities
    ''')
    travel_authorities = cursor.fetchall()

    conn.close()
    return cto_applications, leave_applications, travel_authorities

@app.route('/dashboard')
def dashboard():
    cto_applications = get_all_cto_applications()  # Fetch all CTO applications
    leave_applications = get_all_leave_applications()  # Fetch all leave applications
    travel_authorities = get_all_travel_authorities()  # Fetch all travel authorities
    

    cto_pending = len([app for app in cto_applications if not app.is_approved])
    leave_pending = len([app for app in leave_applications if not app.is_approved])
    travel_pending = len([app for app in travel_authorities if not app.is_approved])



    return render_template('dashboard.html', 
                           total_cto=len(cto_applications),
                           total_leave=len(leave_applications),
                           total_travel=len(travel_authorities),
                           pending_cto=cto_pending,
                           pending_leave=leave_pending,
                           pending_travel=travel_pending,
                           cto_applications=cto_applications,
                           leave_applications=leave_applications,
                           travel_authorities=travel_authorities)


@app.route('/recommender_dashboard')
def recommender_dashboard():
    if 'user_id' not in session or session.get('role') != 'recommender':
        flash('Access denied')
        return redirect(url_for('index'))

    recommender_username = session['username']  # Get the logged-in recommender's username

    # Fetch applications assigned to this recommender
    with sqlite3.connect('documents.db') as conn:
        cto_applications = conn.execute('''SELECT * FROM cto_application 
                                           WHERE recommending_approval = ? AND recommending_approval IS NOT NULL''', 
                                        (recommender_username,)).fetchall()
        leave_applications = conn.execute('''SELECT * FROM leave_application 
                                             WHERE recommending_approval = ? AND recommending_approval IS NOT NULL''', 
                                          (recommender_username,)).fetchall()
        travel_authorities = conn.execute('''SELECT * FROM travel_authority 
                                             WHERE recommending_approval = ? AND recommending_approval IS NOT NULL''', 
                                          (recommender_username,)).fetchall()
    return render_template('recommender_dashboard.html', 
                           cto_applications=cto_applications, 
                           leave_applications=leave_applications, 
                           travel_authorities=travel_authorities)

LEAVE_TYPE_MAP = {
    "1": 'Vacation Leave',
    "2": 'Mandatory/Forced Leave',
    "3": 'Sick Leave',
    "4": 'Maternity Leave',
    "5": 'Paternity Leave',
    "6": 'Special Privilege Leave',
    "7": 'Solo Parent Leave',
    "8": 'Study Leave',
    "9": '10-Day VAWC Leave',
    "10": 'Rehabilitation Privilege',
    "11": 'Speical Leave for Women',
    "12": 'Calamity Leave',
    "13": 'Adoption Leave',
    "14": 'Monetization'
 }

@app.route('/recommended_applications')
def recommended_applications():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    with sqlite3.connect('documents.db') as conn:
        cto_recommended_apps = conn.execute('SELECT * FROM recommended_applications WHERE app_type = "cto"').fetchall()
        leave_recommended_apps = conn.execute('''
            SELECT id, name, position, days, start_date, end_date, leave_type, date_recommended 
            FROM recommended_applications 
            WHERE app_type = "leave"
        ''').fetchall()
 
        # Apply leave type mapping
        leave_recommended_apps = [
            (app[0], app[1], app[2], app[3], app[4], app[5], LEAVE_TYPE_MAP.get(app[6], "Unknown Leave Type"), app[7])
            for app in leave_recommended_apps
        ]

        travel_recommended_apps = conn.execute('SELECT * FROM recommended_applications WHERE app_type = "travel_authority"').fetchall()

    return render_template('recommended_applications.html', cto_recommended_apps=cto_recommended_apps, leave_recommended_apps=leave_recommended_apps, travel_recommended_apps=travel_recommended_apps)

@app.route('/approved_applications')
def approved_applications():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    with sqlite3.connect('documents.db') as conn:
        cto_approved_apps = conn.execute('SELECT * FROM approved_applications WHERE app_type = "cto"').fetchall()
        leave_approved_apps = conn.execute('SELECT * FROM approved_applications WHERE app_type = "leave"').fetchall()
        travel_approved_apps = conn.execute('SELECT * FROM approved_applications WHERE app_type = "travel_authority"').fetchall()

    return render_template('approved_applications.html', cto_approved_apps=cto_approved_apps, leave_approved_apps=leave_approved_apps, travel_approved_apps=travel_approved_apps)

@app.route('/unit_head_dashboard')
def unit_head_dashboard():
    if 'user_id' not in session or session.get('role') != 'unit_head':
        flash('Access denied')
        return redirect(url_for('index'))

    unit_head_username = session['username']  # Ensure the logged-in unit head's username is used

    with sqlite3.connect('documents.db') as conn:
        # Fetch applications specifically assigned to the logged-in unit head
        cto_applications = conn.execute('''
            SELECT id, name, position, days, start_date, end_date 
            FROM cto_application 
            WHERE recommending_approval = ? 
            AND approval_status = "Pending"
        ''', (unit_head_username,)).fetchall()

        leave_applications = conn.execute('''
            SELECT id, name, position, days, start_date, end_date 
            FROM leave_application 
            WHERE recommending_approval = ? 
            AND approval_status = "Pending"
        ''', (unit_head_username,)).fetchall()

        travel_authorities = conn.execute('''
            SELECT id, name, position, purpose, start_date, end_date, destination 
            FROM travel_authority 
            WHERE recommending_approval = ? 
            AND approval_status = "Pending"
        ''', (unit_head_username,)).fetchall()

    return render_template('unit_head_dashboard.html', 
                           cto_applications=cto_applications, 
                           leave_applications=leave_applications, 
                           travel_authorities=travel_authorities)

@app.route('/recommended_head')
def recommended_head():
    if 'user_id' not in session or session.get('role') != 'unit_head':
        flash('Access denied')
        return redirect(url_for('index'))

    unit_head_username = session['username']  # Logged-in unit head's username

    with sqlite3.connect('documents.db') as conn:
        # Fetch recommended CTO applications by this unit head
        cto_recommended_apps = conn.execute('''
            SELECT id, name, position, days, start_date, end_date, date_recommended 
            FROM recommended_applications 
            WHERE app_type = "cto" AND recommending_approval = ?
        ''', (unit_head_username,)).fetchall()

        # Fetch recommended Leave applications by this unit head
        leave_recommended_apps = conn.execute('''
            SELECT id, name, position, days, start_date, end_date, leave_type, date_recommended 
            FROM recommended_applications 
            WHERE app_type = "leave" AND recommending_approval = ?
        ''', (unit_head_username,)).fetchall()

        # Fetch recommended Travel Authority applications by this unit head
        travel_recommended_apps = conn.execute('''
            SELECT id, name, position, purpose, start_date, end_date, destination, date_recommended 
            FROM recommended_applications 
            WHERE app_type = "travel_authority" AND recommending_approval = ?
        ''', (unit_head_username,)).fetchall()
    
    return render_template('recommended_head.html', 
                           cto_recommended_apps=cto_recommended_apps, 
                           leave_recommended_apps=leave_recommended_apps, 
                           travel_recommended_apps=travel_recommended_apps)

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):
    if request.method == 'POST':
        name = request.form['name']
        username = request.form['username']
        position = request.form['position']
        division = request.form['division']  # New field
        office = request.form['office']      # New field
        role = request.form['role']
        
        # Update the user details including the position
        with sqlite3.connect('users.db') as conn:
            conn.execute('UPDATE users SET name = ?, username = ?, position = ?, division = ?, office = ?, role = ? WHERE id = ?', 
                         (name, username, position, division, office, role, user_id))
            conn.commit()

        # If the logged-in user is being edited, update the session position
        if user_id == session.get('user_id'):
            session['position'] = position  # Update the position in session data

        flash('User details updated successfully')
        return redirect(url_for('view_users'))

    # Fetch user details for the form
    with sqlite3.connect('users.db') as conn:
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()

    return render_template('edit_user.html', user=user)

def add_office_and_salary_columns():
    with sqlite3.connect('users.db') as conn:
        # Add 'office' column if it doesn't already exist
        try:
            conn.execute('ALTER TABLE users ADD COLUMN office TEXT')
        except sqlite3.OperationalError:
            print("Column 'office' already exists.")
        
        # Add 'salary' column if it doesn't already exist
        try:
            conn.execute('ALTER TABLE users ADD COLUMN salary INTEGER')
        except sqlite3.OperationalError:
            print("Column 'salary' already exists.")

        # Add 'email' column if it doesn't already exist
        try:
            conn.execute('ALTER TABLE users ADD COLUMN email TEXT')
            print("Column 'email' added to users table.")
        except sqlite3.OperationalError:
            # Column already exists
            pass

        conn.commit()

@app.route('/update_user_info', methods=['POST'])
def update_user_info():
    # Check if user is logged in
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    user_id = session['user_id']

    # Capture the updated information from the form
    office = request.form.get('office')
    salary = request.form.get('salary')
    position = request.form.get('position')
    division = request.form.get('division')
    email = request.form.get('email')  # Capture email from form

    try:
        # Connect to the database and update user information
        with sqlite3.connect('users.db') as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE users 
                SET office = ?, salary = ?, position = ?, division = ?, email = ?
                WHERE id = ?''', (office, salary, position, division, email, user_id))
            conn.commit()

        # Update session variables to reflect the changes immediately in the UI
        session.update({
            'office': office,
            'salary': salary,
            'position': position,
            'division': division,
            'email': email
        })

        flash('User information updated successfully!')

    except Exception as e:
        flash(f'Failed to update user information: {e}')
        return redirect(url_for('user_dashboard'))

    # Fetch updated user details and applications for the dashboard view
    user_details = fetch_user_details(user_id)
    cto_applications, leave_applications, travel_authorities = fetch_applications(user_id)

    return render_template('user_dashboard.html', 
                           user_details=user_details, 
                           cto_applications=cto_applications, 
                           leave_applications=leave_applications, 
                           travel_authorities=travel_authorities)

def fetch_user_details(user_id):
    with sqlite3.connect('users.db') as conn:
        return conn.execute('SELECT office, salary, position, division FROM users WHERE id = ?', (user_id,)).fetchone()

def fetch_applications(user_id):
    with sqlite3.connect('documents.db') as conn:
        cto_applications = conn.execute('SELECT id, name, position, days, start_date, end_date, recommending_approval, approval_status, date_recommended FROM cto_application WHERE user_id = ?', (user_id,)).fetchall()
        leave_applications = conn.execute('SELECT id, name, position, days, start_date, end_date, leave_type, recommending_approval, approval_status, date_recommended FROM leave_application WHERE user_id = ?', (user_id,)).fetchall()
        travel_authorities = conn.execute('SELECT id, name, position, purpose, start_date, end_date, destination, recommending_approval, approval_status FROM travel_authority WHERE user_id = ?', (user_id,)).fetchall()
    
    return cto_applications, leave_applications, travel_authorities


@app.route('/download_application/<string:app_type>/<int:app_id>', methods=['GET'])
def download_application(app_type, app_id):
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))
    base_dir = os.path.join('static', 'generated_files')
    base_dir = os.path.normpath(base_dir)
    if app_type == 'cto':
        file_pattern = os.path.join(base_dir, f'cto_application_{app_id}_*.xlsx')
    elif app_type == 'leave':
        file_pattern = os.path.join(base_dir, f'leave_application_{app_id}_*.xlsx')
    elif app_type == 'travel':
        file_pattern = os.path.join(base_dir, f'travel_authority_{app_id}_*.xlsx')
    else:
        flash('Invalid application type')
        return redirect(url_for('user_dashboard'))
    matching_files = glob.glob(file_pattern)
    if matching_files:
        return send_file(matching_files[0], as_attachment=True)
    else:
        flash('Application file not found')
        return redirect(url_for('user_dashboard'))

# Document Tracker Codes --  # Document Tracker Codes  --  # Document Tracker Codes  --  # Document Tracker Codes  --  # Document Tracker Codes  --  # Document Tracker Codes  
# Document Tracker Codes --  # Document Tracker Codes  --  # Document Tracker Codes  --  # Document Tracker Codes  --  # Document Tracker Codes  --  # Document Tracker Codes 
# Document Tracker Codes --  # Document Tracker Codes  --  # Document Tracker Codes  --  # Document Tracker Codes  --  # Document Tracker Codes  --  # Document Tracker Codes 

def user_exists(username):
    with sqlite3.connect('users.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM users WHERE username = ?", (username,))
        return cursor.fetchone() is not None

def convert_to_localtime(utc_dt):
    """Convert UTC datetime to Manila timezone."""
    return utc_dt.replace(tzinfo=pytz.UTC).astimezone(LOCAL_TIMEZONE)

@app.route('/user_dashboard', methods=['GET', 'POST'])
def user_dashboard():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))
    
    user_id = session['user_id']
    username = session.get('username')
    
    if request.method == 'POST':
        # Update session variables from form data
        session['position'] = request.form['position']
        session['division'] = request.form['division']
        session['office'] = request.form['office']
        session['salary'] = request.form['salary']
        return redirect(url_for('user_dashboard'))

    # Fetch user details and statistics
    user_details, stats = get_user_info_and_stats(user_id)
    
    # Fetch applications (CTO, Leave, Travel Authority) from `documents.db`
    with sqlite3.connect('documents.db') as conn:
        cursor = conn.cursor()
        
        # CTO Applications
        cursor.execute('''SELECT id, name, position, days, start_date, end_date, recommending_approval, 
                          approval_status, rejection_comment, date_recommended 
                          FROM cto_application WHERE user_id = ?''', (user_id,))
        cto_applications = cursor.fetchall()
        
        # Leave Applications
        cursor.execute('''SELECT id, name, position, days, start_date, end_date, leave_type, recommending_approval, 
                          approval_status, rejection_comment, date_recommended 
                          FROM leave_application WHERE user_id = ?''', (user_id,))
        leave_applications = cursor.fetchall()
        
        # Travel Authority Applications
        cursor.execute('''SELECT id, name, position, purpose, start_date, end_date, destination, 
                          recommending_approval, approval_status, rejection_comment 
                          FROM travel_authority WHERE user_id = ?''', (user_id,))
        travel_authorities = cursor.fetchall()
    
    # Fetch submitted and received documents from `document_tracker.db`
    with sqlite3.connect('document_tracker.db') as conn:
        cursor = conn.cursor()
        
        # Documents submitted by the user
        cursor.execute('''
            SELECT id, document_type, forwarded_to, office_forwarded_to, details, submission_date, received 
            FROM documents 
            WHERE submitted_by = ? AND is_archived = 0
        ''', (username,))
        submitted_documents = cursor.fetchall()

        # Fetch documents forwarded to the user (including the 'received' status and comments)
        cursor.execute('''SELECT id, name, position, division, office, document_type, submitted_by, details, received, comments 
                        FROM documents WHERE forwarded_to = ? AND is_archived = 0''', (username,))
        received_documents = cursor.fetchall()

        # --- Build a set of all document IDs to fetch histories for ---
        doc_ids = set(doc[0] for doc in submitted_documents)
        doc_ids.update(doc[0] for doc in received_documents)

        forwarding_histories = {}
        manila_tz = pytz.timezone('Asia/Manila')
        for document_id in doc_ids:
            cursor.execute(
                '''SELECT forwarded_by, forwarded_to, comments, forwarded_at, date_received 
                   FROM forwarding_history 
                   WHERE document_id = ? ORDER BY forwarded_at DESC''', 
                (document_id,)
            )
            history_records = cursor.fetchall()
            converted_history = []
            for record in history_records:
                forwarded_by, forwarded_to, comments, forwarded_at, date_received = record
                utc_time = datetime.fromisoformat(forwarded_at).replace(tzinfo=pytz.utc)
                manila_time = utc_time.astimezone(manila_tz).strftime('%Y-%m-%d %H:%M:%S')
                date_received_str = date_received if date_received else 'N/A'
                converted_history.append((forwarded_by, forwarded_to, comments, manila_time, date_received_str))
            forwarding_histories[document_id] = converted_history

        # Fetch receiving history and convert timestamps to Manila time
        receiving_histories = {}
        cursor.execute("SELECT document_id, received_by, office_received, date_received FROM receiving_history")
        receiving_records = cursor.fetchall()
        for record in receiving_records:
            document_id, received_by, office_received, date_received = record
            if date_received:
                date_received = datetime.fromisoformat(date_received).astimezone(manila_tz).strftime('%Y-%m-%d %H:%M:%S')
            
            if document_id not in receiving_histories:
                receiving_histories[document_id] = []
            receiving_histories[document_id].append((received_by, office_received, date_received))

    # Fetch the office info for each `forwarded_to` entry in submitted_documents
    forwarded_to_offices = {}
    with sqlite3.connect('users.db') as users_conn:
        users_cursor = users_conn.cursor()
        forwarded_to_usernames = [doc[2] for doc in submitted_documents]  # Collect all "forwarded_to" usernames
        placeholders = ','.join('?' * len(forwarded_to_usernames))
        
        if forwarded_to_usernames:  # Only query if there are usernames
            users_cursor.execute(f'''SELECT username, office FROM users WHERE username IN ({placeholders})''', 
                                 forwarded_to_usernames)
            
            # Map each username to their office
            forwarded_to_offices = {row[0]: row[1] for row in users_cursor.fetchall()}

    # Determine office options based on selected division
    division = session.get('division', '')
    officeOptions = {
        'OSDS': ["Accounting", "Admin", "Budget", "ICT", "Legal", "Office of the ASDS", "Office of the SDS", "Payroll", "Personnel", "Records"],
        'SGOD': ["Education Facilities", "Health", "HRD", "Planning and Research", "SGOD", "SMME", "SMN"],
        'CID': ["ALS", "CID", "LR", "PSDS"]
    }.get(division, [])

    # Render the template with all required data
    return render_template(
        'user_dashboard.html',
        user_details=user_details,
        stats=stats,
        cto_applications=cto_applications,
        leave_applications=leave_applications,
        travel_authorities=travel_authorities,
        submitted_documents=submitted_documents,
        received_documents=received_documents,  # <-- now includes comments as doc[9]
        forwarding_histories=forwarding_histories,
        receiving_histories=receiving_histories,
        forwarded_to_offices=forwarded_to_offices,
        officeOptions=officeOptions
    )

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect('document_tracker.db')
    return db

# Close database connection after request
@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

@app.route('/document_tracker')
def document_tracker():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    db = get_db()
    cursor = db.cursor()
    
    # Fetch submitted documents by the user, including office_forwarded_to
    cursor.execute("""
        SELECT id, name, document_type, status, date_submitted, office_forwarded_to 
        FROM documents 
        WHERE submitted_by = ?
    """, (session['user_id'],))
    submitted_documents = cursor.fetchall()

    
    # Fetch forwarded documents to the user
    cursor.execute("""
        SELECT id, name, document_type, status, date_received 
        FROM documents 
        WHERE forwarded_to = ?
    """, (session['user_id'],))
    forwarded_documents = cursor.fetchall()

    return render_template('document_tracker.html', submitted_documents=submitted_documents, forwarded_documents=forwarded_documents)


@app.route('/submit_document_tracker', methods=['POST'])
def submit_document_tracker():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))
    
    user_id = session['user_id']
    username = session.get('username')  # Retrieve username if stored in the session
    
    # Retrieve form data
    name = request.form.get('name')
    position = request.form.get('position')
    division = request.form.get('division')
    office = request.form.get('office')
    document_type = request.form.get('document_type')
    forwarded_to = request.form.get('forwarded_to')
    office_forwarded_to = request.form.get('office_forwarded_to')
    details = request.form.get('details', '')  # Using get to avoid KeyError if not provided
    submission_date = datetime.now().strftime("%m/%d/%Y %H:%M:%S")

    # Check if the 'forwarded_to' user exists in users.db
    try:
        with sqlite3.connect('users.db') as users_conn:
            users_cursor = users_conn.cursor()
            users_cursor.execute("SELECT 1 FROM users WHERE username = ?", (forwarded_to,))
            user_exists = users_cursor.fetchone()

        if not user_exists:
            flash("The 'Forwarded to' user does not exist. Please select from the Suggestions.", "error")
            return redirect(url_for('document_tracker'))  # Redirect to the correct page

        # Save to the document_tracker.db database
        with sqlite3.connect('document_tracker.db') as conn:
            conn.execute(
                '''INSERT INTO documents (name, position, division, office, document_type, forwarded_to, office_forwarded_to, details, submitted_by, submission_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                (name, position, division, office, document_type, forwarded_to, office_forwarded_to, details, username, submission_date)
            )
            conn.commit()

        flash('Document submitted successfully')
        return redirect(url_for('user_dashboard'))

    except sqlite3.Error as e:
        flash(f"Database4 error: {e}", "error")
        return redirect(url_for('document_tracker'))


def get_user_info_and_stats(user_id):
    try:
        with sqlite3.connect('users.db') as conn:
            cursor = conn.cursor()
            user_details = cursor.execute('SELECT office, salary FROM users WHERE id = ?', (user_id,)).fetchone()
            if user_details is None:
                raise ValueError(f"No user found with id {user_id}")

        with sqlite3.connect('documents.db') as conn:
            cursor = conn.cursor()
            # Count of pending, recommended, and approved applications
            pending_count = cursor.execute('SELECT COUNT(*) FROM leave_application WHERE user_id = ? AND approval_status = "Pending"', (user_id,)).fetchone()[0]
            recommended_count = cursor.execute('SELECT COUNT(*) FROM leave_application WHERE user_id = ? AND approval_status = "Recommended"', (user_id,)).fetchone()[0]
            approved_count = cursor.execute('SELECT COUNT(*) FROM leave_application WHERE user_id = ? AND approval_status = "Approved"', (user_id,)).fetchone()[0]
            
            stats = {
                'pending': pending_count,
                'recommended': recommended_count,
                'approved': approved_count
            }

        return user_details, stats

    except sqlite3.Error as e:
        print(f"Database1 error: {e}")
        return None, None
    except Exception as e:
        print(f"Error: {e}")
        return None, None

@app.route('/forward_document_page/<int:document_id>', methods=['GET', 'POST'])
def forward_document_page(document_id):
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))
    
    # Fetch document details from the database
    with sqlite3.connect('document_tracker.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT document_type, name, division, submission_date, details FROM documents WHERE id = ?", (document_id,))
        document_details = cursor.fetchone()
        
    if not document_details:
        flash('Document not found', 'error')
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        forwarded_to = request.form.get('forwarded_to')
        comments = request.form.get('comments')

        # Check if the 'forwarded_to' user exists
        if not user_exists(forwarded_to):
            flash("The 'Forwarded to' user does not exist. Please select from the suggestions.", "error")
            return redirect(url_for('forward_document_page', document_id=document_id))
        
        with sqlite3.connect('document_tracker.db') as conn:
            cursor = conn.cursor()
            cursor.execute('''UPDATE documents SET forwarded_to = ?, comments = ? WHERE id = ?''', (forwarded_to, comments, document_id))
            conn.commit()
        
        flash('Document forwarded successfully')
        return redirect(url_for('user_dashboard'))
    
    return render_template('forward_document.html', document_id=document_id, document_details=document_details)


@app.route('/forward_document', methods=['POST'])
def forward_document():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    document_id = request.form.get('document_id')
    forwarded_to = request.form.get('forwarded_to')
    
    # Check if forwarded_to is retrieved correctly and if user exists
    if not forwarded_to or not user_exists(forwarded_to):
        flash("The 'Forwarded to' user does not exist. Please select from the suggestions.", "error")
        return redirect(url_for('forward_document_page', document_id=document_id))

    return redirect(url_for('forward_document_page', document_id=document_id))


@app.route('/submit_forward_document', methods=['POST'])
def submit_forward_document():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    document_id = request.form['document_id']
    forwarded_to = request.form.get('forwarded_to')  # The new recipient of the document
    forwarded_by = session.get('username')
    comments = request.form.get('comments', '')  # Optional comments for forwarding
    forwarded_at = datetime.now(pytz.timezone('Asia/Manila')).strftime('%Y-%m-%d %H:%M:%S')

    with sqlite3.connect('document_tracker.db') as conn:
        cursor = conn.cursor()

        # Insert into forwarding history
        cursor.execute(
            '''INSERT INTO forwarding_history (document_id, forwarded_by, forwarded_to, comments, forwarded_at)
               VALUES (?, ?, ?, ?, ?)''',
            (document_id, forwarded_by, forwarded_to, comments, forwarded_at)
        )

        # Mark the document as not received for the new recipient
        cursor.execute(
            '''UPDATE documents
               SET received = 0, forwarded_to = ?
               WHERE id = ?''',
            (forwarded_to, document_id)
        )

        conn.commit()

    return redirect(url_for('user_dashboard'))

# Define the relative path to the database
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'document_tracker.db')

@app.route('/delete_document/<int:document_id>', methods=['POST'])
def delete_document(document_id):
    try:
        # Connect to the database
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Update the document entry to mark it as archived
        cursor.execute("UPDATE documents SET is_archived = 1 WHERE id = ?", (document_id,))

        # Commit the transaction if a row was updated
        if cursor.rowcount > 0:
            conn.commit()
            success = True
        else:
            conn.rollback()
            success = False

        # Close the connection
        conn.close()

        return jsonify(success=success)
    
    except Exception as e:
        conn.rollback()
        return jsonify(success=False, error=str(e))


def get_db_connection(db_name='document_tracker.db'):
    conn = sqlite3.connect(db_name)
    conn.row_factory = sqlite3.Row
    return conn

@app.route('/receive_document', methods=['POST'])
def receive_document():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    document_id = request.form['document_id']
    received_by = session.get('username')  # Or get from form data
    office_received = session.get('office')  # Or get from form data
    date_received = datetime.now(pytz.timezone('Asia/Manila')).strftime('%Y-%m-%d %H:%M:%S')  # Use Manila time

    with sqlite3.connect('document_tracker.db') as conn:
        cursor = conn.cursor()

        # Check if the document has already been received
        cursor.execute('''SELECT received FROM documents WHERE id = ?''', (document_id,))
        document = cursor.fetchone()

        if document and document[0] == 1:  # Document has already been received
            return jsonify({"status": "already_received"})

        # Insert receiving record
        cursor.execute(
            '''INSERT INTO receiving_history (document_id, received_by, office_received, date_received)
               VALUES (?, ?, ?, ?)''',
            (document_id, received_by, office_received, date_received)
        )

        # Mark the document as received
        cursor.execute(
            '''UPDATE documents
               SET received = 1
               WHERE id = ?''',
            (document_id,)
        )
        conn.commit()
    return jsonify({"status": "success"})
    
@app.route('/document_history/<int:document_id>')
def document_history(document_id):
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    # Connect to the document tracker database
    with sqlite3.connect('document_tracker.db') as conn:
        cursor = conn.cursor()

        # Fetch document details using the correct column name
        cursor.execute('''SELECT document_type, name, division, submission_date, details
                          FROM documents
                          WHERE id = ?''', (document_id,))
        document_details = cursor.fetchone()

        # Fetch receiving history for the document
        cursor.execute('''SELECT received_by, office_received, date_received
                          FROM receiving_history
                          WHERE document_id = ?
                          ORDER BY date_received DESC''', (document_id,))
        receiving_history = cursor.fetchall()

        # Fetch forwarding history for the document
        cursor.execute('''SELECT forwarded_by, forwarded_to, comments, forwarded_at
                          FROM forwarding_history
                          WHERE document_id = ?
                          ORDER BY forwarded_at DESC''', (document_id,))
        forwarding_history = cursor.fetchall()

    # Render the template with the fetched data
    return render_template(
        'document_history.html',
        document_id=document_id,
        document_details=document_details,
        receiving_history=receiving_history,
        forwarding_history=forwarding_history
    )

@app.route('/print_documents', methods=['GET'])
def print_documents():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))
    
    username = session.get('username')

    # Fetch received documents with the date they were received
    with sqlite3.connect('document_tracker.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''SELECT d.id, d.name, d.position, d.division, d.office, d.document_type, d.details, 
                                 d.submitted_by, rh.date_received
                          FROM documents d
                          JOIN receiving_history rh ON d.id = rh.document_id
                          WHERE rh.received_by = ?''', (username,))
        received_documents = cursor.fetchall()

    # Fetch forwarded documents with the timestamp of when they were forwarded
    with sqlite3.connect('document_tracker.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''SELECT d.id, d.name, d.position, d.division, d.office, d.document_type, d.details, 
                                 fh.forwarded_to, fh.forwarded_at
                          FROM documents d
                          JOIN forwarding_history fh ON d.id = fh.document_id
                          WHERE fh.forwarded_by = ?''', (username,))
        forwarded_documents = cursor.fetchall()

    return render_template(
        'print_documents.html',
        received_documents=received_documents,
        forwarded_documents=forwarded_documents
    )

@app.route('/documents_history')
def documents_history():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    username = session.get('username')
    search_query = request.args.get('search', '')

    with sqlite3.connect('document_tracker.db') as conn:
        cursor = conn.cursor()

        # Fetch receiving history for the logged-in user, including document details
        cursor.execute('''
            SELECT d.document_type, rh.received_by, rh.office_received, rh.date_received, d.submitted_by,
                   fh.forwarded_by, fh.comments, d.details
            FROM receiving_history rh
            JOIN documents d ON rh.document_id = d.id
            LEFT JOIN (
                SELECT document_id, MAX(forwarded_at) as max_forwarded_at
                FROM forwarding_history
                GROUP BY document_id
            ) fh_max ON fh_max.document_id = d.id
            LEFT JOIN forwarding_history fh ON fh.document_id = d.id AND fh.forwarded_at = fh_max.max_forwarded_at
            WHERE rh.received_by = ?
              AND (
                  d.document_type LIKE ?
                  OR d.submitted_by LIKE ?
                  OR rh.received_by LIKE ?
                  OR IFNULL(fh.forwarded_by, '') LIKE ?
                  OR IFNULL(fh.forwarded_to, '') LIKE ?
              )
            ORDER BY rh.date_received DESC
        ''', (
            username,
            f'%{search_query}%',
            f'%{search_query}%',
            f'%{search_query}%',
            f'%{search_query}%',
            f'%{search_query}%'
        ))
        receiving_history = cursor.fetchall()

        # Fetch forwarding history for the logged-in user, including document details
        cursor.execute('''
            SELECT d.document_type, fh.forwarded_by, fh.forwarded_to, fh.comments, fh.forwarded_at, d.submitted_by, d.details
            FROM forwarding_history fh
            JOIN documents d ON fh.document_id = d.id
            WHERE fh.forwarded_by = ?
              AND (
                  d.document_type LIKE ?
                  OR d.submitted_by LIKE ?
                  OR fh.forwarded_to LIKE ?
                  OR fh.forwarded_by LIKE ?
              )
            ORDER BY fh.forwarded_at DESC
        ''', (
            username,
            f'%{search_query}%',
            f'%{search_query}%',
            f'%{search_query}%',
            f'%{search_query}%'
        ))
        forwarding_history = cursor.fetchall()

    return render_template('documents_history.html', 
                           receiving_history=receiving_history, 
                           forwarding_history=forwarding_history, 
                           search_query=search_query)
# BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT 
# BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT 
# BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT BAC PROCUREMENT 

@app.route('/bac_proc', methods=['GET', 'POST'])
def bac_proc():
    if request.method == 'POST':
        # Handle form submission
        bac_document_type = request.form['bac_document_type']
        bac_details = request.form['bac_details']
        supplier = request.form.get('supplier')
        # Save data to the database
        flash("Document submitted successfully!", "success")
        return redirect(url_for('bac_proc'))
    return render_template('bac_proc.html')

@app.route('/submit_bac_proc', methods=['POST'])
def submit_bac_proc():
    # Logic for handling the form submission
    pass

@app.route('/clear_documents_db', methods=['POST'])
def clear_documents_db():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied')
        return redirect(url_for('index'))

    try:
        with sqlite3.connect('documents.db') as conn:
            conn.execute('DELETE FROM cto_application')
            conn.execute('DELETE FROM leave_application')
            conn.execute('DELETE FROM travel_authority')
            conn.execute('DELETE FROM recommended_applications')
            conn.execute('DELETE FROM approved_applications')
            conn.commit()
        flash('documents.db cleared successfully', 'success')
    except Exception as e:
        flash(f'Error clearing documents.db: {e}', 'danger')

    return redirect(url_for('admin_dashboard'))

@app.route('/clear_document_tracker_db', methods=['POST'])
def clear_document_tracker_db():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied')
        return redirect(url_for('index'))

    try:
        with sqlite3.connect('document_tracker.db') as conn:
            conn.execute('DELETE FROM documents')
            conn.execute('DELETE FROM forwarding_history')
            conn.execute('DELETE FROM receiving_history')
            conn.commit()
        flash('document_tracker.db cleared successfully', 'success')
    except Exception as e:
        flash(f'Error clearing document_tracker.db: {e}', 'danger')

    return redirect(url_for('admin_dashboard'))

@app.route('/service_record')
def service_record():
    if 'user_id' not in session:
        flash('Please log in first')
        return redirect(url_for('index'))

    user_id = session['user_id']
    username = session.get('username')

    # Fetch service record details from the database
    with sqlite3.connect('users.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''SELECT position, division, office, salary 
                          FROM users 
                          WHERE id = ?''', (user_id,))
        user_details = cursor.fetchone()

    # Check if user details exist
    if not user_details:
        flash('Service record not found', 'danger')
        return redirect(url_for('user_dashboard'))

    return render_template('service_record.html', user_details=user_details, username=username)



def ensure_email_column():
    with sqlite3.connect('users.db') as conn:
        try:
            conn.execute('ALTER TABLE users ADD COLUMN email TEXT')
            print("Column 'email' added to users table.")
        except sqlite3.OperationalError:
            # Column already exists
            pass

if __name__ == '__main__':
    ensure_email_column()  # Ensure email column exists before running the app
    port = int(os.environ.get('PORT', 5001))  # Get the PORT from environment, default to 5000
    app.run(debug=True, host='0.0.0.0', port=port)